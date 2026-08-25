#!/usr/bin/env python3
"""
Convert a legacy .xls (BIFF) workbook to .xlsx so the normal openpyxl
pipeline can edit it.

Why this exists:
  - openpyxl cannot open .xls at all (InvalidFileException).
  - pandas read_excel().to_excel() is NOT a converter: it silently drops
    every sheet except the first, plus styles, merged cells and number
    formats. Never use it for conversion.
  - This script copies every sheet cell-by-cell via xlrd, preserving all
    sheets, values (dates/booleans/errors typed correctly), merged cell
    ranges, and basic formatting (font name/size/bold/italic/color, solid
    fills, number formats, alignment, column widths, row heights).

Formula recovery (best-effort, BIFF8 only):
  xlrd's public API exposes only cached formula results, but its internal
  decompiler (xlrd.formula.decompile_formula) can turn the raw RPN token
  stream back into formula text. We scan the BIFF stream ourselves,
  collect FORMULA (0x0006) and SHRFMLA (0x04BC) records, resolve shared-
  formula tExp pointers against their templates, and write recovered
  formulas as live '=...' cells. Cells whose RPN cannot be decompiled
  fall back to their cached constant value — check `formulas_fallback`
  in the JSON output and rebuild those by hand.

Output JSON shape (mirrors recalc.py / verify_workbook.py):
    {
      "status": "success" | "error",
      "output": "path/to/out.xlsx",
      "sheets": ["名称", ...],
      "cells_copied": int,
      "formulas_recovered": int,   # written as live formulas
      "formulas_fallback": int,    # kept as cached constants — rebuild!
      "fallback_cells": [          # present when formulas_fallback > 0
        {"cell": "Sheet!C2", "cached_value": ..., "reason": "..."}, ...
      ],
      "formatting": "preserved" | "values_only",
      "note": "..."
    }

Usage:
    python scripts/xls_to_xlsx.py <input.xls> [output.xlsx]
"""

import json
import re
import struct
import sys
from pathlib import Path

OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

_HALIGN = {1: "left", 2: "center", 3: "right", 4: "fill", 5: "justify"}
_VALIGN = {0: "top", 1: "center", 2: "bottom"}

# collapses degenerate ranges like F8:F8 (xlrd renders single-cell 3D refs
# as ranges) so the output reads naturally
_DEGENERATE_RANGE = re.compile(r"(\$?[A-Z]{1,3}\$?[0-9]+):(\$?[A-Z]{1,3}\$?[0-9]+)")


def _sniff(path: Path) -> str | None:
    """Return an error message if the file is not a real BIFF .xls."""
    head = path.open("rb").read(512)
    if head.startswith(OLE_MAGIC):
        return None
    lower = head.lstrip()[:200].lower()
    if lower.startswith(b"<") or b"<html" in lower or b"<table" in lower:
        return (
            "This file is an HTML table saved with a .xls extension "
            "(common ERP/web export), not a real Excel file. "
            "Read it with pandas: pd.read_html(path) — do not use xlrd."
        )
    if head.startswith(b"PK"):
        return (
            "This file is a ZIP-based workbook (.xlsx/.xlsm) with a .xls "
            "extension. Rename it and open it with openpyxl directly."
        )
    return "Not a BIFF .xls file (unknown magic bytes)."


def _workbook_stream(path: Path):
    """Return the raw BIFF Workbook stream, correctly sliced.

    locate_named_stream returns (mem, base, size). For streams in the
    main FAT, `mem` is the ENTIRE file buffer and the stream lives at
    mem[base:base+size] — scanning from 0 walks OLE header bytes and
    yields garbage (this was a real bug: formula count silently came out
    0 for any workbook stream > 4096 bytes). Always slice.
    """
    from xlrd.compdoc import CompDoc

    cd = CompDoc(path.read_bytes())
    for name in ("Workbook", "Book"):
        try:
            mem, base, size = cd.locate_named_stream(name)
        except Exception:
            continue
        if mem:
            return mem[base : base + size]
    return None


def _scan_records(stream):
    """Yield (opcode, payload, worksheet_index) for every BIFF record.
    worksheet_index counts worksheet BOF substreams (dt=0x0010) so
    formula records can be attributed to their sheet."""
    pos, sheet_idx = 0, -1
    while pos + 4 <= len(stream):
        op, ln = struct.unpack("<HH", stream[pos : pos + 4])
        payload = stream[pos + 4 : pos + 4 + ln]
        if op == 0x0809 and len(payload) >= 4:
            if struct.unpack("<H", payload[2:4])[0] == 0x0010:
                sheet_idx += 1
        yield op, payload, sheet_idx
        pos += 4 + ln


def _normalize(txt: str) -> str:
    def collapse(m):
        return m.group(1) if m.group(1) == m.group(2) else m.group(0)

    return _DEGENERATE_RANGE.sub(collapse, txt)


def _extract_formulas(path: Path, bk):
    """Best-effort recovery of formula text from the BIFF stream.

    Returns ({(sheet_idx, rowx, colx): "=formula"}, failed) where failed
    is a list of (sheet_idx, rowx, colx, reason) for cells whose RPN
    could not be decompiled, or None when the BIFF stream could not be
    scanned at all (pre-BIFF8 file / corrupt OLE directory).
    """
    if getattr(bk, "biff_version", 0) != 80:
        return {}, None

    from xlrd import formula as xf

    stream = _workbook_stream(path)
    if stream is None:
        return {}, None

    # xlrd bug workaround: for 3D range refs (e.g. SUMIF(Sheet!A:A,...)),
    # decompile_formula calls get_cell_range_addr WITHOUT forwarding
    # browx/bcolx, and adjust_cell_addr_biff8 crashes on `rowx -= browx`
    # with None. Relative-encoded refs need the CURRENT cell's coords as
    # the base (substituting 0 silently shifts every ref), so we wrap the
    # function and inject them from a holder set per cell below.
    holder: dict = {}
    orig_range = xf.get_cell_range_addr

    def patched_range(data, pos, bv, reldelta, browx=None, bcolx=None):
        if browx is None:
            browx = holder.get("browx")
        if bcolx is None:
            bcolx = holder.get("bcolx")
        return orig_range(data, pos, bv, reldelta, browx, bcolx)

    shared: dict = {}   # (sheet_idx, anchor_row, anchor_col) -> template rpn
    cells: list = []    # (sheet_idx, rowx, colx, rpn)
    try:
        for op, payload, si in _scan_records(stream):
            if si < 0:
                continue
            if op == 0x0006 and len(payload) >= 22:  # FORMULA
                rowx, colx = struct.unpack("<HH", payload[0:4])
                fl = struct.unpack("<H", payload[20:22])[0]
                cells.append((si, rowx, colx, payload[22 : 22 + fl]))
            elif op == 0x04BC and len(payload) >= 10:  # SHRFMLA
                rw1 = struct.unpack("<H", payload[0:2])[0]
                c1 = payload[4]
                fl = struct.unpack("<H", payload[8:10])[0]
                shared[(si, rw1, c1)] = payload[10 : 10 + fl]
    except Exception:
        return {}, None

    formulas: dict = {}
    failed: list = []  # (sheet_idx, rowx, colx, reason)
    xf.get_cell_range_addr = patched_range
    try:
        for si, rowx, colx, rpn in cells:
            holder["browx"], holder["bcolx"] = rowx, colx
            try:
                if rpn and rpn[0] == 0x01 and len(rpn) >= 5:  # tExp → shared
                    arow, acol = struct.unpack("<HH", rpn[1:5])
                    template = shared.get((si, arow, acol))
                    if template is None:
                        failed.append((si, rowx, colx, "shared template not found"))
                        continue
                    txt = xf.decompile_formula(
                        bk, template, len(template), xf.FMLA_TYPE_SHARED,
                        browx=rowx, bcolx=colx, blah=0, r1c1=0,
                    )
                else:
                    txt = xf.decompile_formula(
                        bk, rpn, len(rpn), xf.FMLA_TYPE_CELL,
                        browx=rowx, bcolx=colx, blah=0, r1c1=0,
                    )
                if not txt or txt.startswith("SHARED FMLA"):
                    failed.append((si, rowx, colx, "decompiler returned placeholder"))
                    continue
                formulas[(si, rowx, colx)] = "=" + _normalize(txt)
            except Exception as e:
                failed.append((si, rowx, colx, f"{type(e).__name__}: {e}"))
    finally:
        xf.get_cell_range_addr = orig_range
    return formulas, failed


def _rgb(colour_map, idx) -> str | None:
    rgb = colour_map.get(idx)
    if not rgb:
        return None
    return "%02X%02X%02X" % rgb


def _make_style_cache(bk):
    """Map xf_index -> dict of openpyxl style objects, built lazily."""
    from openpyxl.styles import Alignment, Font, PatternFill

    cache: dict = {}

    def styles_for(xf_index: int) -> dict:
        if xf_index in cache:
            return cache[xf_index]
        out: dict = {}
        try:
            xf = bk.xf_list[xf_index]
            f = bk.font_list[xf.font_index]
            color = _rgb(bk.colour_map, f.colour_index)
            out["font"] = Font(
                name=f.name or None,
                size=(f.height / 20.0) if f.height else None,
                bold=f.weight >= 700,
                italic=bool(f.italic),
                color=color,
            )
            if xf.background.fill_pattern == 1:  # solid
                bg = _rgb(bk.colour_map, xf.background.pattern_colour_index)
                if bg and bg != "FFFFFF":
                    out["fill"] = PatternFill("solid", start_color=bg)
            fmt = bk.format_map.get(xf.format_key)
            if fmt and fmt.format_str and fmt.format_str.lower() != "general":
                out["number_format"] = fmt.format_str
            ha = _HALIGN.get(xf.alignment.hor_align)
            va = _VALIGN.get(xf.alignment.vert_align)
            if ha or va:
                out["alignment"] = Alignment(horizontal=ha, vertical=va)
        except Exception:
            out = {}
        cache[xf_index] = out
        return out

    return styles_for


def convert(src: str, dst: str | None = None) -> dict:
    path = Path(src)
    if not path.exists():
        return {"status": "error", "error": f"File {src} does not exist"}

    sniff_err = _sniff(path)
    if sniff_err:
        return {"status": "error", "error": sniff_err}

    try:
        import xlrd
    except ImportError:
        return {"status": "error", "error": "xlrd is required (pip install xlrd)"}
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    out = Path(dst) if dst else path.with_suffix(".xlsx")

    # formatting_info=True yields XF records (fonts/fills/formats) and
    # merged ranges; some files make xlrd choke on it, so fall back.
    has_fmt = True
    try:
        bk = xlrd.open_workbook(path, formatting_info=True)
    except Exception:
        bk = xlrd.open_workbook(path)
        has_fmt = False

    styles_for = _make_style_cache(bk) if has_fmt else None
    formulas, f_failed = _extract_formulas(path, bk)

    wb = Workbook()
    wb.remove(wb.active)
    cells = 0
    recovered = 0
    for si, name in enumerate(bk.sheet_names()):
        sh = bk.sheet_by_name(name)
        ws = wb.create_sheet(name)
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                t = sh.cell_type(r, c)
                if t in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    continue
                v = sh.cell_value(r, c)
                # LibreOffice stores booleans as =TRUE()/=FALSE() formula
                # records; keep the typed boolean, not a pseudo-formula.
                fx = None
                if t != xlrd.XL_CELL_BOOLEAN:
                    fx = formulas.get((si, r, c))
                if fx is not None:
                    v = fx
                    recovered += 1
                elif t == xlrd.XL_CELL_DATE:
                    v = xlrd.xldate.xldate_as_datetime(v, bk.datemode)
                elif t == xlrd.XL_CELL_BOOLEAN:
                    v = bool(v)
                elif t == xlrd.XL_CELL_ERROR:
                    v = xlrd.biffh.error_text_from_code.get(v, "#ERR")
                cell = ws.cell(r + 1, c + 1, v)
                cells += 1
                if styles_for is not None:
                    st = styles_for(sh.cell_xf_index(r, c))
                    if "font" in st:
                        cell.font = st["font"]
                    if "fill" in st:
                        cell.fill = st["fill"]
                    if "alignment" in st:
                        cell.alignment = st["alignment"]
                    if "number_format" in st and t != xlrd.XL_CELL_DATE:
                        cell.number_format = st["number_format"]
        for rlo, rhi, clo, chi in getattr(sh, "merged_cells", []):
            ws.merge_cells(
                start_row=rlo + 1, end_row=rhi,
                start_column=clo + 1, end_column=chi,
            )
        if has_fmt:
            for cidx, info in getattr(sh, "colinfo_map", {}).items():
                if info.width:
                    ws.column_dimensions[get_column_letter(cidx + 1)].width = (
                        info.width / 256.0
                    )
            for ridx, info in getattr(sh, "rowinfo_map", {}).items():
                if info.height and getattr(info, "height_mismatch", 0):
                    ws.row_dimensions[ridx + 1].height = info.height / 20.0

    # Recovered formula cells carry no cached value; make every viewer
    # (Excel/WPS/Numbers) recompute on open.
    wb.calculation.fullCalcOnLoad = True
    wb.save(out)

    result = {
        "status": "success",
        "output": str(out),
        "sheets": bk.sheet_names(),
        "cells_copied": cells,
        "formulas_recovered": recovered,
        "formulas_fallback": 0,
        "formatting": "preserved" if has_fmt else "values_only",
        "note": (
            "All sheets, values and basic formatting copied. Recovered "
            "formulas are written as live '=...' cells (recomputed on "
            "open). Verify with the QA pipeline (recalc.py)."
        ),
    }
    if f_failed is None:
        result["formulas_fallback"] = "unknown"
        result["note"] = (
            "All sheets, values and basic formatting copied. Formula "
            "recovery was skipped (pre-BIFF8 file or unscannable stream) — "
            "any formulas are now cached constants. Inspect the source and "
            "rebuild needed totals/ratios as live formulas."
        )
    elif f_failed:
        from openpyxl.utils import get_column_letter as _gcl

        names = bk.sheet_names()
        result["formulas_fallback"] = len(f_failed)
        result["fallback_cells"] = [
            {
                "cell": f"{names[si]}!{_gcl(colx + 1)}{rowx + 1}",
                "cached_value": bk.sheet_by_index(si).cell_value(rowx, colx),
                "reason": reason,
            }
            for si, rowx, colx, reason in f_failed[:50]
        ]
        result["note"] += (
            f" {len(f_failed)} formula(s) could not be decompiled and "
            "remain cached constants — rebuild each cell listed in "
            "fallback_cells as a live formula."
        )
    return result


def main() -> None:
    # Windows cmd/PowerShell consoles often default to a legacy codepage
    # (e.g. GBK); printing CJK sheet names in the JSON would raise
    # UnicodeEncodeError. Reconfigure stdout to UTF-8 when possible.
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
    if len(sys.argv) < 2:
        print(
            "Usage: python xls_to_xlsx.py <input.xls> [output.xlsx]\n\n"
            "Converts legacy .xls to .xlsx: all sheets, typed values, merged\n"
            "cells, basic formatting, and best-effort formula recovery\n"
            "(check formulas_fallback for cells needing manual rebuild).\n"
            "Default output: same path with .xlsx suffix."
        )
        sys.exit(1)
    dst = sys.argv[2] if len(sys.argv) > 2 else None
    print(json.dumps(convert(sys.argv[1], dst), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
